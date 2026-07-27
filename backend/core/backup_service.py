import io
import json
import os
import re
import shutil
import sqlite3
import tempfile
import uuid
import zipfile
from decimal import Decimal
from datetime import date, datetime, timedelta
from pathlib import Path

from django.apps import apps
from django.conf import settings as django_settings
from django.db import connections
from django.db.models.fields.files import FieldFile
from django.utils import timezone

SKIP_MODEL_LABELS = {('sessions', 'session')}


def filename_stamp():
    return timezone.now().strftime('%Y%m%d_%H%M%S')


def sheet_title(model):
    raw = f'{model._meta.app_label}_{model.__name__}'
    raw = re.sub(r'[\[\]\*\/\\\?\:]', '_', raw)
    return (raw[:31] or 'data')


def cell_value(val):
    if val is None:
        return ''
    if isinstance(val, FieldFile):
        try:
            return val.name or ''
        except Exception:
            return str(val) if val else ''
    if isinstance(val, uuid.UUID):
        return str(val)
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, timedelta):
        return str(val)
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False)
    if isinstance(val, bytes):
        return val.hex()
    return str(val)


RANGE_FIELD_CANDIDATES = [
    'created_at',
    'date',
    'order_date',
    'sale_date',
    'quotation_date',
    'purchase_date',
    'payment_date',
    'job_date',
    'transaction_date',
    'rent_date',
    'month',
    'join_date',
    'loan_date',
    'date_given',
    'deduction_date',
    'updated_at',
]


def _normalize_range_bounds(start_date=None, end_date=None):
    if start_date and isinstance(start_date, datetime):
        start_dt = start_date
    elif start_date:
        start_dt = datetime.combine(start_date, datetime.min.time())
    else:
        start_dt = None

    if end_date and isinstance(end_date, datetime):
        end_dt = end_date
    elif end_date:
        end_dt = datetime.combine(end_date, datetime.max.time())
    else:
        end_dt = None

    if start_dt and timezone.is_naive(start_dt):
        start_dt = timezone.make_aware(start_dt, timezone.get_current_timezone())
    if end_dt and timezone.is_naive(end_dt):
        end_dt = timezone.make_aware(end_dt, timezone.get_current_timezone())

    return start_dt, end_dt


def _pick_range_field(model):
    fields = {f.name: f for f in model._meta.concrete_fields}
    for name in RANGE_FIELD_CANDIDATES:
        field = fields.get(name)
        if field is not None:
            return field
    return None


def _apply_date_range(qs, model, start_date=None, end_date=None):
    if not start_date and not end_date:
        return qs

    field = _pick_range_field(model)
    if field is None:
        return qs

    lookup = field.name
    filters = {}
    if getattr(field, 'get_internal_type', lambda: '')() == 'DateTimeField':
        start_dt, end_dt = _normalize_range_bounds(start_date, end_date)
        if start_dt:
            filters[f'{lookup}__gte'] = start_dt
        if end_dt:
            filters[f'{lookup}__lte'] = end_dt
    else:
        if start_date:
            filters[f'{lookup}__gte'] = start_date
        if end_date:
            filters[f'{lookup}__lte'] = end_date

    return qs.filter(**filters)


def build_excel_bytes(start_date=None, end_date=None):
    """
    Export every non-proxy model: concrete columns plus comma-separated related PKs for each
    local many-to-many. Uses a standard (non-write_only) workbook so Excel reliably contains data.
    """
    from openpyxl import Workbook

    wb = Workbook()
    used_titles = {}
    first_sheet = True

    for model in apps.get_models():
        if model._meta.proxy:
            continue
        label = (model._meta.app_label, model._meta.model_name)
        if label in SKIP_MODEL_LABELS:
            continue

        title = sheet_title(model)
        base_title = title
        n = 2
        while title in used_titles:
            suffix = f'_{n}'
            title = (base_title[: max(1, 31 - len(suffix))] + suffix)
            n += 1
        used_titles[title] = True

        if first_sheet:
            ws = wb.active
            ws.title = title
            first_sheet = False
        else:
            ws = wb.create_sheet(title=title)

        concrete = list(model._meta.concrete_fields)
        m2m_local = list(model._meta.local_many_to_many)

        headers = [f.name for f in concrete]
        for m in m2m_local:
            headers.append(f'm2m_{m.name}')

        ws.append(headers)

        qs = _apply_date_range(model.objects.all(), model, start_date=start_date, end_date=end_date)
        prefetch = [m.name for m in m2m_local]
        if prefetch:
            qs = qs.prefetch_related(*prefetch)

        try:
            rows = qs.iterator(chunk_size=500)
        except Exception:
            rows = qs

        for obj in rows:
            row = []
            for field in concrete:
                try:
                    v = field.value_from_object(obj)
                except Exception:
                    v = getattr(obj, field.name, None)
                row.append(cell_value(v))
            for m in m2m_local:
                try:
                    pks = list(getattr(obj, m.name).values_list('pk', flat=True))
                    row.append(','.join(str(x) for x in pks))
                except Exception:
                    row.append('')
            ws.append(row)

    if not used_titles:
        ws = wb.create_sheet(title='info')
        ws.append(['message'])
        ws.append(['No models exported.'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_sql_bytes():
    engine = django_settings.DATABASES['default']['ENGINE']
    if 'sqlite' not in engine:
        raise RuntimeError('SQL text backup is only supported for SQLite databases.')
    db_path = Path(django_settings.DATABASES['default']['NAME'])
    if not db_path.is_file():
        raise RuntimeError('Database file not found.')

    fd, tmp_path = tempfile.mkstemp(suffix='.sqlite3')
    os.close(fd)
    try:
        # Hot-copy live DB with SQLite backup API (works with WAL; plain copy can miss pages)
        src = sqlite3.connect(str(db_path))
        dst_live = sqlite3.connect(tmp_path)
        try:
            src.backup(dst_live)
        finally:
            dst_live.close()
            src.close()

        conn = sqlite3.connect(tmp_path)
        try:
            out = io.BytesIO()
            for line in conn.iterdump():
                out.write(line.encode('utf-8'))
                out.write(b'\n')
        finally:
            conn.close()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
    out.seek(0)
    return out


def restore_sql_bytes(sql_bytes):
    engine = django_settings.DATABASES['default']['ENGINE']
    if 'sqlite' not in engine:
        raise RuntimeError('Restore is only supported for SQLite databases.')

    db_path = Path(django_settings.DATABASES['default']['NAME'])
    if not db_path.parent.exists():
        raise RuntimeError('Database directory not found.')

    fd, tmp_path = tempfile.mkstemp(suffix='.sqlite3')
    os.close(fd)
    try:
        sql_text = sql_bytes.decode('utf-8')
        temp_conn = sqlite3.connect(tmp_path)
        try:
            temp_conn.executescript(sql_text)
            temp_conn.commit()
        finally:
            temp_conn.close()

        connections.close_all()

        src = sqlite3.connect(tmp_path)
        dst = sqlite3.connect(str(db_path))
        try:
            src.backup(dst)
            dst.commit()
        finally:
            dst.close()
            src.close()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def restore_media_tree(extracted_media_root: Path):
    if not extracted_media_root.exists():
        return

    media_root = Path(django_settings.MEDIA_ROOT)
    media_root.mkdir(parents=True, exist_ok=True)

    for child in list(media_root.iterdir()):
        if child.name == 'backups':
            continue
        if child.is_dir():
            shutil.rmtree(child, ignore_errors=True)
        else:
            try:
                child.unlink()
            except OSError:
                pass

    for path in extracted_media_root.rglob('*'):
        if not path.is_file():
            continue
        rel = path.relative_to(extracted_media_root)
        dest = media_root / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, dest)


def _should_skip_media_path(full_path: Path, media_root: Path) -> bool:
    """Avoid nesting backup zips and huge auto-backup folders."""
    try:
        rel = full_path.relative_to(media_root)
    except ValueError:
        return True
    parts = rel.parts
    if len(parts) >= 1 and parts[0] == 'backups':
        return True
    return False


def add_uploaded_media_to_zip(zf: zipfile.ZipFile):
    """
    Append files under MEDIA_ROOT as media/... so logos and uploads are restorable beside DB/SQL.
    Skips media/backups/** to avoid recursive zip bloat.
    """
    root = Path(django_settings.MEDIA_ROOT)
    if not root.is_dir():
        return

    for path in root.rglob('*'):
        if not path.is_file():
            continue
        if _should_skip_media_path(path, root):
            continue
        try:
            arcname = Path('media') / path.relative_to(root)
            zf.write(path, arcname.as_posix())
        except OSError:
            continue


def prune_old_backups(directory: Path, keep=40):
    files = sorted(directory.glob('backup_auto_*'), key=lambda p: p.stat().st_mtime, reverse=True)
    for p in files[keep:]:
        try:
            p.unlink()
        except OSError:
            pass


def write_automatic_backup_bundle(settings_row):
    if not settings_row.backup_include_excel and not settings_row.backup_include_sql:
        return
    dest = Path(django_settings.MEDIA_ROOT) / 'backups' / 'auto'
    dest.mkdir(parents=True, exist_ok=True)
    stamp = filename_stamp()
    excel_buf = None
    sql_buf = None
    if settings_row.backup_include_excel:
        excel_buf = build_excel_bytes()
    if settings_row.backup_include_sql:
        sql_buf = build_sql_bytes()

    if excel_buf and sql_buf:
        zip_path = dest / f'backup_auto_{stamp}.zip'
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f'system_backup_{stamp}.xlsx', excel_buf.getvalue())
            zf.writestr(f'system_backup_{stamp}.sql', sql_buf.getvalue())
            add_uploaded_media_to_zip(zf)
    elif excel_buf:
        (dest / f'backup_auto_{stamp}.xlsx').write_bytes(excel_buf.getvalue())
    elif sql_buf:
        (dest / f'backup_auto_{stamp}.sql').write_bytes(sql_buf.getvalue())

    prune_old_backups(dest, keep=40)
